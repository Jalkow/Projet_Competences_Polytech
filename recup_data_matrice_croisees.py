from openpyxl import load_workbook
import re
import os
import json

def get_color(cell_color, wb):
    """
    Retourne la couleur RGB sous forme de chaîne RRGGBB même si c'est une couleur de thème.
    """
    if cell_color is None:
        return None

    # Couleur RGB directe
    if cell_color.type == 'rgb' and cell_color.rgb:
        return cell_color.rgb[-6:]  # ignore le FF alpha

    # Couleur basée sur un thème
    if cell_color.type == 'theme' and cell_color.theme is not None:
        # wb._theme est le fichier XML du thème, openpyxl peut le résoudre
        # on peut approximer avec wb.theme_colors pour récupérer la couleur
        try:
            theme_colors = wb.theme_colors
            theme_rgb = theme_colors[cell_color.theme]
            if theme_rgb:
                return theme_rgb[-6:]
        except Exception:
            return None

    # Couleur automatique (noir)
    if cell_color.type == 'auto':
        return "000000"

    return None

def get_liaison_type(cell):
    if not cell.font or not cell.font.color:
        return None

    color = cell.font.color
    if color is None:
        return None

    if(color.type == "theme"):
        # vert 
        if(color.theme == 9):
            return "ciblee"

    if(color.type == "rgb"):
        # bleu
        if(color.rgb == 'FF0070C0'):
            return "fournie"
        
        # violet
        elif (color.rgb == 'FF7030A0'):
            return "variant"

    return None

def extract_infos_matrices(matrice_file_path, output_folder_path):

    excel_workbook = load_workbook(matrice_file_path)

    result = {
        "liste_UE":[],
        "liens": []
    }

    is_first_sheet = True
    for sheet in excel_workbook.worksheets:

        ue_by_col_number = {}
        regular_expression_ue_code = re.compile(r"UE\d+")
        regular_expression_CE_AC = r"^(CE|AC)\s\d+\.\d+$" # commence par CE ou AC, suivi d'une espace puis d'un truc de la forme 1.1

        # Détection des colonnes UE (max_row car nom des UE devraient être vers le haut du excel donc ça évite de regarder le reste pour rien)
        for row in sheet.iter_rows(max_row=15):
            for cell in row:
                # Search renvoie un Match (true) si trouvé sinon un NULL (false)
                if isinstance(cell.value, str) and regular_expression_ue_code.search(cell.value):
                    # clé utilisée plus tard pour trouver si il y a liaison avec CE / AC
                    ue_by_col_number[cell.column] = cell.value
                    
                    # faire la liste des UE lors de la lecture de la premiere feuille excel
                    if(is_first_sheet):
                        separated_ue_text = cell.value.split()
                        ue_code = separated_ue_text[0]
                        ue_name = " ".join(separated_ue_text[1:])
                        result["liste_UE"].append({
                            "code" : ue_code,
                            "nom" : ue_name, 
                        })

        # Parcours des lignes pour avoir liaisons UE avec CE et AC
        for row in sheet.iter_rows():

            code = None
            description = None

            # row est la liste des cellules de la ligne 
            # pour avoir num de la ligne on prend numero de ligne de la premiere cellule
            row_number = row[0].row

            for cell in row:
                if isinstance(cell.value, str):

                    if re.match(regular_expression_CE_AC, cell.value):
                        code = cell.value

                    elif code and description is None:
                        description = cell.value

            # Si pas une ligne avec une composante essentielle ou apprentissage critique
            if not code or not description:
                continue

            infos = {
                "code": code,
                "nom": description,
                "UE_liees": {}
            }

            # col = clé (numero col) et ue = valeur (code UE + nom)
            for col, ue in ue_by_col_number.items():

                separated_ue_text = ue.split()
                ue_code = separated_ue_text[0]

                cell = sheet.cell(row_number, col)

                if str(cell.value).lower() == "x":

                    # Récupère le type de la liaison (fournie, ciblee ou variant) ou None si il n'y a pas de liaison
                    type_liaison = get_liaison_type(cell)

                    if type_liaison is not None:
                        infos["UE_liees"][ue_code] = {"type_liaison" : type_liaison}

            result["liens"].append(infos)
        

        is_first_sheet = False

    # création du fichier json avec les données scrappées
    output_file_path = os.path.join(output_folder_path, "matrice_output.json")
    with open(output_file_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=4)

    return result